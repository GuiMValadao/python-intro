# Capítulo 21 - Trabalhando com arquivos excel
# Este capítulo introduz o módulo openpyxl que pode ser usado para trabalhar
# com arquivos Excel.
# ----------------------------------------------------------
# Arquivos Excel
# Apesar de arquivos csv serem um modo convenientes e simples de lidar com
# dados, é comum precisar escrever ou ler arquivos Excel diretamente. Para
# isso, existem várias bibliotecas em Python. Uma amplamente usada é a
# OpenPyXL, originalmente escrita para oferecer suporte a arquivos Excel
# 2010. É um projeto de código aberto e bem documentado.
# A OpenPyXL possui recursos para:
#   * ler e escrever workbooks do Excel
#   * criar/acessar planilhas do Excel
#   * criar fórmulas do Excel
#   * criar gráficos (com suporte de módulos adicionais)
# -------------------------------------------------------
# A classe Workbook do OpenPyxl
# O elemento principal da biblioteca OpenPyXL é a classe Workbook. Ela
# pode ser importada do módulo:
from openpyxl import Workbook

# Uma nova instância (na memória) de Workbook pode ser criada usando a classe
# Workbook.
wb = Workbook()
# -------------------------------------------------
# Objetos planilha do Openpyxl
# Um workbook é, sempre, criado com pelo menos uma planilha. Você pode
# pegar a planillha atualmente ativa usando a propriedade Workbook.active:
ws = wb.active
# Você pode criar planilhas adicionais usando o método do workbook create_sheet():
ws = wb.create_sheet("Mysheet")
# Pode acessar ou atualizar o título da planilha usando a propriedade title:
ws.title = "Novo Titulo"
# A cor de fundo da aba guardando o título é branca por padrão. Pode ser
# alterada fornecendo uma cor RRGGBB para worksheet.sheet_properties.tabColor:
ws.sheet_properties.tabColor = "1072BA"
# -------------------------------------------
# Trabalhando com células
# É possível acessar as células dentro de uma planilha. Uma célula pode ser
# acessada diretamente como chaves na planilha, por exemplo:
ws["A1"] = 42
# ou
cell = ws["A1"]
# Isto retorna um objeto célula; você pode obter o valor da célula usando
# a propriedade value, por exemplo:
print(cell.value)
# Também existe o método worksheet.cell() que fornece acesso a células usando
# a notação de linha e coluna:
d = ws.cell(row=4, column=2, value=10)
# Uma linha de valores também pode ser acrescentada à posição atual dentro do
# arquivo Excel usando append:
ws.append([1, 2, 3])
# Ela acrescentará uma linha ao arquivo Excel contendo 1, 2 e 3.
# Grupos de células podem ser acessados usando fatiamento:
cell_range = ws["A1":"C2"]
# Grupos de linhas ou colunas também podem ser obtidos:
col = ws["C"]
col_range = ws["C:D"]
row10 = ws[10]
row_range = ws[5:10]

# O valor de uma célula também pode ser uma fórmula Excel como:
ws["A3"] = "=SUM(A1, A2)"
# Um workbook é, na verdade, apenas uma estrutura na memória; precisa ser
# salvo em um arquivo para armazenamento permanente. Estes workbooks
# podem ser salvos usando o método save(). Este método pega um nome de arquivo
# e escreve o Workbook no formato Excel.
# wb.save("balances.xlsx")
# -----------------------------------------------------------
# Aplicação de exemplo de criação de arquivo excel
# A seguinte aplicação simples cria um Workbook com duas planilhas. Também
# contém uma fórmula Excel simples que soma os valores guardados nas outras
# células:
###############################################
# from openpyxl import Workbook


# def main():
#     print("Iniciando escrita de exemplo Excel com openPyXL")
#     workbook = Workbook()
#     ws = workbook.active
#     ws.title = "minha planilha"
#     ws.sheet_properties.tabColor = "1072BA"

#     ws["A1"] = 42
#     ws["A2"] = 12
#     ws["A3"] = "=SUM(A1, A2)"
#     ws2 = workbook.create_sheet(title="outra planilha")
#     ws2["A1"] = 3.42
#     ws2.append([1, 2, 3])
#     ws2.cell(column=2, row=1, value=15)

#     workbook.save("exemplo.xlsx")
#     print("Escrita do exemplo excel finalizada")


# if __name__ == "__main__":
#     main()
###############################################
# Carregando um workbook de um arquivo excel
# Obviamente, em muitos casos é necesspario não apenas criar arquivos
# Excel para exportar dados como importar dados de um arquivo Excel existente.
# Isto pode ser feito usando a função load_workbook() do OpenPyXL. Ela abre
# o arquivo Excel especificado (no modo apenas leitura por padrão) e retorna
# um objeto Workbook.
# from openpyxl import load_workbook
# workbook = load_workbook(filename='exemplo.xlsx')
# Agora você pode acessar uma lista de planilhas, seus nomes, obter a planilha
# atualmente ativa etc usando propriedades fornecidas pelo objeto workbook
# A seguinte aplicação simples lê o arquivo Excel criado acima:
# from openpyxl import load_workbook


# def main():
#     print("Iniciando leitura de arquivo Excel usando openPyXL")
#     workbook = load_workbook(filename="exemplo.xlsx")
#     print(workbook.active)
#     print(workbook.sheetnames)
#     print(workbook.worksheets)
#     print("-" * 10)
#     ws = workbook["minha planilha"]
#     print(ws["A1"])
#     print(ws["A1"].value)
#     print(ws["A2"].value)
#     print(ws["A3"].value)

#     print("-" * 10)
#     for sheet in workbook:
#         print(sheet.title)
#     print("-" * 10)
#     cell_range = ws["A1":"A3"]
#     for cell in cell_range:
#         print(cell[0].value)
#     print("-" * 10)
#     print("Leitura de arquivo Excel com openPyXL finalizada.")


# if __name__ == "__main__":
#     main()
